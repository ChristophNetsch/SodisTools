#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jul 16 17:26:24 2020

@author: ml
"""
from trello import TrelloClient
#from styleframe import StyleFrame
import os
 #from pathlib import Path
#import pandas as pd
from datetime import datetime
#import shutil
#import slack
#from slack.errors import SlackApiError
#import numpy as np
import configparser as ConfigParser
#from datetime import datetime
#from collections import Counter, OrderedDict
import openpyxl
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#Trello authentication keys
def connect_to_trello(KeyFilePath):
    #not in use
    keyParser = ConfigParser.RawConfigParser()   
    keyParser.read(KeyFilePath)
    
    api_key=keyParser.get("Trello Access","API Key").replace(" ", "")
    api_secret=keyParser.get("Trello Access","API Secret").replace(" ", "")
    
    client_trello = TrelloClient(
        api_key=api_key,
        api_secret=api_secret
    )
    return client_trello

def connect_to_trello_env(enviroment_list):
    try:
        api_key = os.environ['API_KEY']
        api_secret = os.environ['API_SECRET']
        
        client_trello = TrelloClient(
            api_key=api_key,
            api_secret=api_secret
        )
        
    except:
        print('ERROR: Report aborted. Please provide Trello and Slack credentials in env.list-file.')
    return client_trello


client = connect_to_trello(r"keys/trello_access.txt")
#client = connect_to_trello_env(r"keys/env.list")

config_name = os.listdir(r"config_folder/")
configFilePath = []

for i in range(0,len(config_name),1):
    name = (r"config_folder/" + config_name[i])
    name2 = "config_folder/.DS_Store"
    if name != name2:
        configFilePath.append(r"config_folder/" + config_name[i])
    else:
        ii = i 
if ii:
    del(config_name[ii])

def cal_excel (config_name,configFilePath):
    #AUSLESEN DER KONFIGDATEI
    configParser = ConfigParser.RawConfigParser()   
    configParser.read(configFilePath)
    board_id = configParser.get("Section IDs","Boards-IDs").split(",")
    alist_id = configParser.get("Section IDs","Active List-IDs").split(",")
    sflist_id = configParser.get("Section IDs","Semi-Final List-IDs").split(",")
    flist_id = configParser.get("Section IDs","Final List-IDs").split(",")
    list_id_order = configParser.get("Section IDs","List-ID Forward Order").split(",")
    #report_repeat = configParser.get("Parameters","repeat")
    report_name = configParser.get("Parameters","report_name")
        #exclude spaces " "
    
    def exclude_space_in_list_entries(lst):
        for i in range(0,len(lst),1):
            lst[i] = lst[i].replace(" ", "")
        return lst
    exclude_space_in_list_entries(board_id)
    exclude_space_in_list_entries(alist_id)
    exclude_space_in_list_entries(sflist_id)
    exclude_space_in_list_entries(flist_id)
    exclude_space_in_list_entries(list_id_order)

    #INITIALISIERUNG
    time_now = datetime.now().astimezone()
    
    def get_lists_from_ids(ids_list):
        list_object = []
        #get_name = "get_" + object_type
        for id_counter in range (0,len(ids_list),1):
            list_object.append(client.get_list(ids_list[id_counter]))
        return list_object
    
    def get_boards_from_ids(ids_list):
        list_object = []
        #get_name = "get_" + object_type
        for id_counter in range (0,len(ids_list),1):
            list_object.append(client.get_board(ids_list[id_counter]))
        return list_object
    
    def get_members_from_board_ids():
        member_dict = {}
        
        for org in client.list_organizations():
            org_mem = org.get_members()
            for i in range(0,len(org_mem),1):
                (firstWord, *_) = org_mem[i].full_name.split(maxsplit=1)
                member_dict[org_mem[i].id] = {
                    "First Name": firstWord,
                    'Name':org_mem[i].full_name,
                    "Username":org_mem[i].username,
                    'ID':org_mem[i].id
                    }    
        return member_dict

    def get_cards_from_list_id (list_ids):
        tmp_cards=[]
    
        for tmp_board in client.list_boards():
            if tmp_board.id in board_id:            
               for tmp_list in tmp_board.all_lists(): 
                    if tmp_list.id in list_ids:
                        tmp_cards+=tmp_list.list_cards()
        return tmp_cards
    
    def cal_active_cards (alist_id,sflist_id,flist_id):
        tmp_cards = get_cards_from_list_id (alist_id)
        r_a_cards = len(tmp_cards)
        tmp_cards += get_cards_from_list_id (sflist_id)
        r_sf_cards = len(tmp_cards) - r_a_cards
        tmp_cards += get_cards_from_list_id (flist_id)
        r_f_cards = len(tmp_cards) - r_sf_cards - r_a_cards
        r_cards = len(tmp_cards)
        return r_a_cards, r_sf_cards, r_f_cards, r_cards, tmp_cards

    #RUN
    #alist_ob = get_lists_from_ids(alist_id)
    #flist_ob = get_lists_from_ids(flist_id)
    #sflist_ob = get_lists_from_ids(sflist_id)
    #list_ob_order = get_lists_from_ids(list_id_order)
    
    member_dict = get_members_from_board_ids()
    
    r_a_cards, r_sf_cards, r_f_cards, r_cards, tmp_cards = cal_active_cards (alist_id,sflist_id,flist_id)

    def write_excel():
        def check_if_already_exist(card):
            result = False
            for j in range(1,sheet.max_row,1):
                if sheet["J"+ str(j)].value == card.id:
                    result = True
                    break
            return result, j
        
        def move_dims(card,num):
            dim = sheet.dimensions
            if sheet.max_row >= num:
                dim_col = dim.replace("A1:","")
                dim_new = "A"+ str(num) + ":" + dim_col
                sheet.move_range((dim_new), rows=1, cols=0)
                
        def write_line(card,num,update=False):
            if (update == False) or sheet["A"+ str(num)].value != card.name:
                sheet["A"+ str(num)].value = card.name
            
            lst = client.get_list(card.idList)
            if (update == False) or sheet["B"+ str(num)].value != lst.name:
                sheet["B"+ str(num)].value = lst.name
            
            due = card.due_date
            if not due:
                if (update == False) or sheet["C"+ str(num)].value != "❌":
                    sheet["C"+ str(num)].value = "❌"
            else:
                if (update == False) or sheet["C"+ str(num)].value != card.due_date.date():
                    sheet["C"+ str(num)].value = card.due_date.date()
                    sheet["C"+ str(num)].value
                
            if card.is_due_complete:                
                if (update == False) or sheet["D"+ str(num)].value != "✅":
                    sheet["D"+ str(num)].value = "✅"
            else: 
                if (update == False) or sheet["D"+ str(num)].value != "❌":
                    sheet["D"+ str(num)].value = "❌"
            
            if (update == False) or sheet["E"+ str(num)].value != card.date_last_activity.date():
                sheet["E"+ str(num)].value = card.date_last_activity.date()

            m_nms = ""
            m_ids = card.member_ids
            for m in m_ids:
                m_nms += member_dict[m]["Name"] + ", "
            if m_nms == "":
                if (update == False) or sheet["F"+ str(num)].value != "❌":
                    sheet["F"+ str(num)].value = "❌"
            else:
                if (update == False) or sheet["F"+ str(num)].value != m_nms[:-2]:
                    sheet["F"+ str(num)].value = m_nms[:-2]
            
            cmts = card.comments
            c_str = ""
            for cmt in cmts:
                r_tag = {tag.strip("#") for tag in cmt["data"]["text"].split() if tag.startswith("#")}
                if r_tag:
                    c_str += "- " + cmt["data"]["text"] + "; "
            if (update == False) or sheet["G"+ str(num)].value != c_str:
                sheet["G"+ str(num)].value = c_str
            
            if (update == False) or sheet["H"+ str(num)].value != card.description:
                sheet["H"+ str(num)].value = card.description
            
            if (update == False) or sheet["I"+ str(num)].value != card.short_url:
                sheet["I"+ str(num)].value = card.short_url
            
            if (update == False) or sheet["J"+ str(num)].value != card.id:
                sheet["J"+ str(num)].value = card.id

        #RUN            
        c=0
        for s in wb.worksheets:
            if s.title in config_name :
                sheet = s
                c=1
        if c == 0:
            sheet = wb.copy_worksheet(wb.worksheets[0])
            sheet.title = config_name[0:30]

            sheet["A1"] = "Chronik"
            sheet["C1"] =report_name
            br_nm = ""
            for b_id in board_id:
                brd = client.get_board(b_id)
                br_nm += brd.name
            sheet["F1"] = br_nm
            sheet["H1"] = "Letzte Aktualisierung"
            sheet["J1"] = time_now
            
            sheet["A2"] = "Kartenname"
            sheet["B2"] = "Liste"
            sheet["C2"] = "Frist"
            sheet["D2"] = "Frist eingehalten"
            sheet["E2"] = "Letzte Aktivität"
            sheet["F2"] = "Verantwortung"
            sheet["G2"] = "#Schlüssel-Kommentar"
            sheet["H2"] = "Beschreibung"
            sheet["I2"] = "Link"
            sheet["J2"] = "Karten ID"   

        for i in range(len(tmp_cards)-1,0,-1):            
            num = 3
            card = tmp_cards[i]
            
            r_check, row_num = check_if_already_exist(card) 
            if r_check == False:
                move_dims(card,num)
                write_line(card,num,update = False)
                wb.save('trello_chronicle.xlsx')

            else:
                write_line(card,row_num,update = True)
                wb.save('trello_chronicle.xlsx')

        
    write_excel()

#Run Calculation
wb = openpyxl.load_workbook('trello_chronicle.xlsx')
for k in range(0,len(config_name),1):
    cal_excel(config_name[k],configFilePath[k])
    print("- updated", k+1 ,"of",len(config_name),"reports in trello_chronicle file -")

wb.save('trello_chronicle.xlsx')


#CREATE CRONIC-EXCEL
def get_card_data(card):
    def _check_data(entry):
        if type(entry)==list:
            string=''
            for sub_entry in entry:
                string+=str(sub_entry)+', \n'
        else:
            string=entry
        return str(string)
    
    def _translate_member_ids(member_ids):
        entry=''
        for member_id in member_ids:
            entry+=card.client.get_member(member_id).full_name + ', \n'
        return entry
    
    def _merge_checklists(checklists):
        merged_lists=[]
        for checklist in checklists:
            for item in checklist.items:
                merged_lists.append(item['name'])
        return merged_lists